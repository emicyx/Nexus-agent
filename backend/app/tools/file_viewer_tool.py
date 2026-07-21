"""文件查看工具：读取多种格式的文件内容。"""
import csv
import io
import json
import logging
from pathlib import Path
from typing import Any

from crewai.tools import BaseTool
from pydantic import BaseModel, Field

from app.tools._file_utils import resolve_read_path, truncate

logger = logging.getLogger("file_viewer")

# 纯文本扩展名
_TEXT_EXTS = {
    ".txt", ".md", ".py", ".js", ".ts", ".java", ".go", ".rs", ".sh",
    ".sql", ".html", ".css", ".yaml", ".yml", ".json", ".csv", ".xml",
    ".ini", ".conf", ".toml", ".log", ".env", ".txt",
}


class FileViewerInput(BaseModel):
    file_path: str = Field(
        ...,
        description=(
            "文件路径。支持绝对路径和相对 /app/data/ 的路径。"
            "例如：'outputs/hello.py' 或 '/app/data/outputs/report.docx'"
        ),
    )


class FileViewerTool(BaseTool):
    """查看文件内容，支持纯文本、Word、Excel、PDF 等格式。"""
    name: str = "view_file"
    description: str = (
        "查看文件内容。支持：纯文本(.txt/.md/.py/.js/.json/.csv/.yaml/.html 等)、"
        "Word(.docx)、Excel(.xlsx)、PDF(.pdf)。大文件自动截断到 8000 字符。"
        "触发时机：需要查看已有文件内容、读取生成的文档、检查代码文件时使用。"
    )
    args_schema: type[BaseModel] = FileViewerInput

    def _run(self, file_path: str = "", **kwargs: Any) -> str:
        if not file_path:
            return "错误：file_path 不能为空"

        path = resolve_read_path(file_path.strip())
        if not path.exists():
            return f"文件不存在: {file_path}"
        if not path.is_file():
            return f"路径不是文件: {file_path}"

        ext = path.suffix.lower()
        try:
            if ext == ".docx":
                text = _read_docx(path)
            elif ext == ".xlsx":
                text = _read_xlsx(path)
            elif ext == ".pdf":
                text = _read_pdf(path)
            elif ext in _TEXT_EXTS:
                text = path.read_text(encoding="utf-8")
            else:
                text = path.read_text(encoding="utf-8")
        except Exception as e:
            logger.exception("file_viewer_error")
            return f"读取文件失败: {e}"

        return truncate(text)


def _read_docx(path: Path) -> str:
    """提取 Word .docx 文本。"""
    from docx import Document

    doc = Document(str(path))
    lines = []
    for para in doc.paragraphs:
        if para.text.strip():
            lines.append(para.text)
    return "\n".join(lines) if lines else "(空文档)"


def _read_xlsx(path: Path) -> str:
    """提取 Excel .xlsx 数据。"""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    out_lines = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        out_lines.append(f"[工作表: {ws_name}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            out_lines.append(" | ".join(cells))
        out_lines.append("")
    wb.close()
    return "\n".join(out_lines) if out_lines else "(空表格)"


def _read_pdf(path: Path) -> str:
    """提取 PDF 文本。"""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        pages.append(f"--- 第 {i + 1} 页 ---\n{text}")
    return "\n\n".join(pages) if pages else "(空PDF)"
